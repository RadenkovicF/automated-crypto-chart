import requests
from requests import Session
from requests.exceptions import ConnectionError, Timeout, TooManyRedirects
import json
from datetime import datetime
import openpyxl
from web3 import Web3

# hier Pfad zum excel Dokument eintragen
#TODO: Excel Path eintragen
path_to_excel = "/crypto_geld.xlsx"
# Wallet adressen
#TODO: Wallet Addressen setzen
mm_address = ""
ledger_address = ""


#für preise benötigt, kostenlos bis 333 anfrangen am Tag
#TODO: Coinmarketcapapi erstellen und key einsetzen
coinmarketcapapi_key = ""
#TODO: binance api key erstellen und einsetzen
binance_api_key = ""
#für etherscan benötigt kostenlos, alternative direkt account bei etherscan. Gleiche konditionen
#TODO: Infura account (Ethereum) erstellen und key einsetzen
infura_key = ""
ethereum_api_url = "https://mainnet.infura.io/v3/" + infura_key

datum = datetime.now()
jahr = datum.strftime("%Y")
monat = datum.strftime("%B")
datum_string = str(monat) + " " + str(jahr)

datum_excel = datum.date()
uhrzeit = datum.time()
uhrzeit_excel = uhrzeit.replace(minute=0, second=0, microsecond=0)



# kostenlose APIs um Preise auf der jeweiligen Chain abzufragen
url_eth = "https://api.etherscan.io/api"
url_polygon = "https://api.polygonscan.com/api"
url_bsc = "https://api.bscscan.com/api"
url_avax = "https://api.snowtrace.io/api"
url_moon = "https://api-moonriver.moonscan.io/api"

# contract address, findet sich in jeder doch
gohm_contract_address = '0x0ab87046fBb341D058F17CBC4c1133F25a20a52f'  # gOhm
sklima_contract_address = "0xb0C22d8D350C67420f06F48936654f567C73E8C8"  # sKlima
smeta_contract_address = "0x09f33EC33052Cd253Db79fFA883E9c12Eb578309"  # sMeta
stime_contract_address = "0x136acd46c134e8269052c62a67042d6bdedde3c9"  # Memo
srome_contract_address = "0x89f52002e544585b42f8c7cf557609ca4c8ce12a"  # sRome


# Hier die coinmarketcap id der Token eintragen. Mit skript "getIdFromSymbol.py"
#TODO: Mit skript "get_id_by_symbol.py" die id des gewünschten Tokens holen
my_assets = {
    #"ETH": "1027",
    "sOhm": "12903",
    "Meta": "14705",
    "Klima": "12873",
    "TIME": "11585",
    #"Strong": "6511",
    "ROME": "15398"
}
# coinmarketapi für bestimmte Preise
url = 'https://pro-api.coinmarketcap.com/v1/cryptocurrency/quotes/latest'
# parameters für die Abfrage mit der conmarketcap API
# Preise deafult in USD. Wenn USD gewünscht, einfach EUR zu USD wechseln oder vielleicht einfach löschen
# zu testen
#TODO: parameter einstellen, alle ID eintragen und Währung auswählen
parameters = {
    'convert': 'EUR',
    "id": "1027,12903,14705,12873,11585,6511,15398"
}

# header für die Abfrage mit coinmarket api
# @API_KEY
headers = {
    'Accepts': 'application/json',
    'X-CMC_PRO_API_KEY': coinmarketcapapi_key,
}

# session coinmarketcap api
session = Session()
session.headers.update(headers)

# Rome (moon)
w3 = Web3(Web3.HTTPProvider("https://rpc.moonriver.moonbeam.network"))
API_ENDPOINT = url_moon + "?module=contract&action=getabi&address=" + str(srome_contract_address)
r = requests.get(url=API_ENDPOINT)
response = r.json()
instance = w3.eth.contract(
    address=Web3.toChecksumAddress(srome_contract_address),
    abi=response["result"]
)
# Holen der amount vom wallet
#TODO: ADDRESSE ÄNDERN
rome_amount = instance.functions.balanceOf(mm_address).call()
# convert to decimal
rome_amount_dec = rome_amount / 1000000000

# Ohm (ERC)
w3 = Web3(Web3.HTTPProvider(ethereum_api_url))
API_ENDPOINT = url_eth + "?module=contract&action=getabi&address=" + str(gohm_contract_address)
r = requests.get(url=API_ENDPOINT)
response = r.json()

instance = w3.eth.contract(
    address=Web3.toChecksumAddress(gohm_contract_address),
    abi=response["result"]
)
# Holen der amount vom wallet
#TODO: ADDRESSE ÄNDERN und ggf. gohm sohm beachten
ohm_index = instance.functions.index().call()
# convert to decimal and convert gOhm to sOhm
sOhm_converted_amount_dec = (ohm_index / 1000000000) * 0.120203816111679969

# Klima (Polygon)
w3 = Web3(Web3.HTTPProvider("https://rpc-mainnet.maticvigil.com"))
API_ENDPOINT = url_polygon + "?module=contract&action=getabi&address=" + str(sklima_contract_address)
r = requests.get(url=API_ENDPOINT)
response = r.json()

instance = w3.eth.contract(
    address=Web3.toChecksumAddress(sklima_contract_address),
    abi=response["result"]
)
# Holen der amount vom wallet
#TODO: ADDRESSE ÄNDERN
klima_amount = instance.functions.balanceOf(ledger_address).call()
# convert to decimal
klima_amount_dec = klima_amount / 1000000000

# Meta (BSC)
w3 = Web3(Web3.HTTPProvider("https://bsc-dataseed1.binance.org"))
API_ENDPOINT = url_bsc + "?module=contract&action=getabi&address=" + str(
    smeta_contract_address) + "&apikey=" + binance_api_key
r = requests.get(url=API_ENDPOINT)
response = r.json()

instance = w3.eth.contract(
    address=Web3.toChecksumAddress(smeta_contract_address),
    abi=response["result"]
)
# Holen der amount vom wallet
#TODO: ADDRESSE ÄNDERN
meta_amount = instance.functions.balanceOf(mm_address).call()
# convert to decimal
meta_amount_dec = meta_amount / 1000000000

# TIME (Avax)
w3 = Web3(Web3.HTTPProvider("https://api.avax.network/ext/bc/C/rpc"))
API_ENDPOINT = url_avax + "?module=contract&action=getabi&address=" + str(stime_contract_address)
r = requests.get(url=API_ENDPOINT)
response = r.json()

instance = w3.eth.contract(
    address=Web3.toChecksumAddress(stime_contract_address),
    abi=response["result"]
)
# Holen der amount vom wallet
#TODO: ADDRESSE ÄNDERN
time_amount = instance.functions.balanceOf(ledger_address).call()
# convert to decimal
time_amount_dec = time_amount / 1000000000


try:
    gesamt = 0
    response = session.get(url, params=parameters)
    data = json.loads(response.text)

    # openpyxl zum öffnen der excel datei
    wb = openpyxl.load_workbook(path_to_excel)
    # ws ist die aktive Seite, diese muss mit dem Namen gematched sein (unten in excel)
    ws = wb["Uebersicht"]
    # mit ws["xy"] werden die zellen der aktiven (ws) Seite beschrieben
    ws["E1"] = datum_excel
    ws["F1"] = uhrzeit_excel


    # sOhm
    current_price = int(data["data"][my_assets["sOhm"]]["quote"]["EUR"]["price"])
    sOhm_val = current_price * sOhm_converted_amount_dec
    gesamt += sOhm_val
    print(data["data"][my_assets["sOhm"]]["symbol"] + ":\t" + "{:.2f}".format(sOhm_val) + "\t@\t" + str(current_price))

    # META
    current_price = int(data["data"][my_assets["Meta"]]["quote"]["EUR"]["price"])
    meta_val = current_price * meta_amount_dec
    gesamt += meta_val
    print(data["data"][my_assets["Meta"]]["symbol"] + ":\t" + "{:.2f}".format(meta_val) + "\t@\t" + str(current_price))

    # Klima
    current_price = int(data["data"][my_assets["Klima"]]["quote"]["EUR"]["price"])
    klima_val = current_price * klima_amount_dec
    gesamt += klima_val
    print(
        data["data"][my_assets["Klima"]]["symbol"] + ":\t" + "{:.2f}".format(klima_val) + "\t@\t" + str(current_price))

    # TIME
    current_price = int(data["data"][my_assets["TIME"]]["quote"]["EUR"]["price"])
    time_val = current_price * time_amount_dec
    gesamt += time_val
    print(data["data"][my_assets["TIME"]]["symbol"] + ":\t" + "{:.2f}".format(time_val) + "\t@\t" + str(current_price))

    # ROME
    current_price = int(data["data"][my_assets["ROME"]]["quote"]["EUR"]["price"])
    rome_val = current_price * rome_amount_dec
    gesamt += rome_val
    print(data["data"][my_assets["ROME"]]["symbol"] + ":\t" + "{:.2f}".format(rome_val) + "\t@\t" + str(current_price))

    print(gesamt)

    # Für die Uebersicht Seite
    ws["G3"] = sOhm_val
    ws["G4"] = meta_val
    ws["G5"] = klima_val
    ws["G6"] = time_val
    ws["G9"] = rome_val
    # zwischen Speichern
    wb.save(path_to_excel)

    # bearbeiten der nächsten Seite (Monatsübersicht)
    # If Seite mit namen monat+jahr existiert noch nicht, dann erstelle die Seite direkt mit den Überschriften
    if datum_string not in wb.sheetnames:
        ws = wb.create_sheet(datum_string)
        ws = wb[datum_string]
        ws["C1"] = "Ohm"
        ws["D1"] = "Klima"
        ws["E1"] = "TIME"
        ws["F1"] = "Meta"
        ws["G1"] = "Rome"

    # wähle die neue bzw. bereits erstellte Seite aus
    ws = wb[datum_string]

    # funktioniert, aber ignoriert viele mögliche fehler
    # Gehe jede zeile (row) der ersten Spalte durch und suche die nächste freie zelle.
    # Die nächste Freie muss im normalfall immer unter der vorherigen sein
    for col in ws.iter_cols(min_row=2, max_col=1, max_row=750):  #
        for cell in col:
            if cell.value is None:
                # falls die Zelle leer ist, füge Datum für die übersicht hinzu
                cell.value = datum_excel
                current_row = cell.row
                print("Datum erstellt")
                break

    # vorbereiten der Felder
    uhrzeit_row = "B" + str(current_row)
    ohm_row = "C" + str(current_row)
    klima_row = "D" + str(current_row)
    time_row = "E" + str(current_row)
    meta_row = "F" + str(current_row)
    rome_row = "G" + str(current_row)

    # formatieren der Felder in Euro und eintragen
    ws[uhrzeit_row] = uhrzeit_excel
    ws[ohm_row].number_format = '#,##0.00€'
    ws[klima_row].number_format = '#,##0.00€'
    ws[time_row].number_format = '#,##0.00€'
    ws[meta_row].number_format = '#,##0.00€'
    ws[rome_row].number_format = '#,##0.00€'

    ws[ohm_row] = sOhm_val
    ws[klima_row] = klima_val
    ws[time_row] = time_val
    ws[meta_row] = meta_val
    ws[rome_row] = rome_val

    # speichern des Dokuments
    wb.save(path_to_excel)

except (ConnectionError, Timeout, TooManyRedirects) as e:
    print(e)
